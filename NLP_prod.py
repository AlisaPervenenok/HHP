import re
from string import punctuation
from array import *
import openpyxl
import xlrd
import numpy as np
from collections import defaultdict
from collections import Counter

class Porter:
	PERFECTIVEGROUND =  re.compile(u"((ив|ивши|ившись|ыв|ывши|ывшись)|((?<=[ая])(в|вши|вшись)))$")
	REFLEXIVE = re.compile(u"(с[яь])$")
	ADJECTIVE = re.compile(u"(ее|ие|ые|ое|ими|ыми|ей|ий|ый|ой|ем|им|ым|ом|его|ого|ему|ому|их|ых|ую|юю|ая|яя|ою|ею)$")
	PARTICIPLE = re.compile(u"((ивш|ывш|ующ)|((?<=[ая])(ем|нн|вш|ющ|щ)))$")
	VERB = re.compile(u"((ила|ыла|ена|ейте|уйте|ите|или|ыли|ей|уй|ил|ыл|им|ым|ен|ило|ыло|ено|ят|ует|уют|ит|ыт|ены|ить|ыть|ишь|ую|ю)|((?<=[ая])(ла|на|ете|йте|ли|й|л|ем|н|ло|но|ет|ют|ны|ть|ешь|нно)))$")
	NOUN = re.compile(u"(а|ев|ов|ие|ье|е|иями|ями|ами|еи|ии|и|ией|ей|ой|ий|й|иям|ям|ием|ем|ам|ом|о|у|ах|иях|ях|ы|ь|ию|ью|ю|ия|ья|я)$")
	RVRE = re.compile(u"^(.*?[аеиоуыэюя])(.*)$")
	DERIVATIONAL = re.compile(u".*[^аеиоуыэюя]+[аеиоуыэюя].*ость?$")
	DER = re.compile(u"ость?$")
	SUPERLATIVE = re.compile(u"(ейше|ейш)$")
	I = re.compile(u"и$")
	P = re.compile(u"ь$")
	NN = re.compile(u"нн$")

	def stem(word):
		word = word.lower()
		word = word.replace(u'ё', u'е')
		m = re.match(Porter.RVRE, word)
		if m and m.groups():
			pre = m.group(1)
			rv = m.group(2)
			temp = Porter.PERFECTIVEGROUND.sub('', rv, 1)
			if temp == rv:
				rv = Porter.REFLEXIVE.sub('', rv, 1)
				temp = Porter.ADJECTIVE.sub('', rv, 1)
				if temp != rv:
					rv = temp
					rv = Porter.PARTICIPLE.sub('', rv, 1)
				else:
					temp = Porter.VERB.sub('', rv, 1)
					if temp == rv:
						rv = Porter.NOUN.sub('', rv, 1)
					else:
						rv = temp
			else:
				rv = temp
			
			rv = Porter.I.sub('', rv, 1)

			if re.match(Porter.DERIVATIONAL, rv):
				rv = Porter.DER.sub('', rv, 1)

			temp = Porter.P.sub('', rv, 1)
			if temp == rv:
				rv = Porter.SUPERLATIVE.sub('', rv, 1)
				rv = Porter.NN.sub(u'н', rv, 1)
			else:
				rv = temp
			word = pre+rv
		return word
	
	#Porter fo list
	def stem_list(list_mess):
		b1 = []
		list_mess.sort()
		for x in list_mess:
			b1.append(Porter.stem(x))
		b1.sort()
		return b1
		
	stem=staticmethod(stem)

class Matcher_topics:
  stopwords = [a for a in punctuation]
  stopwords.append('«')
  stopwords.append('»')
  stopwords2 = ['в', 'не', 'по', 'на', 'прош', 'и', 'при', 'с', 'добр', 'ден', 'для', 'к', 'нет', 'эт', 'как', 'из', 'о', 'у', 'а', 'от', ' ', 'ил', 'во', 'он', 'что', 'то', 'так', 'но', 'спасиб',]
  rb = xlrd.open_workbook('content/Data0406_.xls', formatting_info=True)
  sheet = rb.sheet_by_index(0)
  a = []
  b = []
  c = []
  d1 = [] #сообщения
  d2 = [] #ключи
  d3 = []
  keys = []
  mess = []

#Формирование ключей из заявок за 3 мес
  keys2 = np.array([[],[]], float)
  keys3 = np.array([[],[]], float)
  keys4 = np.array([[],[]], float)
  mess2 = [] #список с заявками за 3 мес по нужному ИТ-реш
  mess3 = [] #список заявок ВСЕГО за три мес, поступивших на ИТ-реш
  mess_one = []

  def __init__(self):
      k = 0
      for rownum in range(self.sheet.nrows):
        row = self.sheet.row_values(rownum)
        for c_el in row:
          self.a.extend(c_el.split())
          self.mess_one.extend(c_el.split())
        self.mess2.append(self.mess_one)
        k = k + 1
        self.mess_one = []
      for y in self.mess2:
        self.d3.append(Porter.stem_list(y))
      self.mess2.clear()
      self.mess2.extend(self.d3)
      self.d3.clear()

      self.a.sort()

      for slovo in self.a:
        self.b.append(Porter.stem(slovo))
      # СТОП СЛОВА
      for stopw2 in self.stopwords2:
        while self.b.count(stopw2) != 0:
          self.b.remove(stopw2)

      #Формирование ключей и весов

      i = 1
      j = 0
      wb_ = openpyxl.Workbook()
      wb_.create_sheet(title = 'Test', index=0)
      sheet_ = wb_['Test']

      for x in self.b:
        if self.c.count(x) == 0 and self.b.count(x) > 2 and len(x) > 1:
          self.c.append(x)

          value = str(self.b.count(x))
          cell = sheet_.cell(row=i, column=1)
          cell.value = value

          value = str(self.mess2[1].count(x))
          sheet_.cell(row=i, column=3)
          cell.value = value

          value = str(x)
          cell = sheet_.cell(row=i, column=4)
          cell.value = value
          i = i + 1

      wb_.save('content/SD2020.xlsx')

      #создать словаль
      word_counts = defaultdict(int) #int() возвращает 0
      for word in self.b:
        word_counts[word] += 1

      word_counts2 = {key: val for key, val in word_counts.items() if val > 1}

      myList = word_counts2.items()
      myList = sorted(myList)

      #Лучший вариант подсчета количества появлений слов
      for y in self.mess2: #список сообщений
        word_counts = Counter(y)

      keys2 = np.zeros([(len(self.mess2)+1),(len(self.c)+3)],float)

      d1 = []
      j = 3
      for y in self.mess2: #список сообщений
        j = j + 1
        i = 1
        sum = 0

        for x in self.c: #список всех стем
          self.d2.append(y.count(x))
          sum += y.count(x)

          keys2[j-4, i-1] = y.count(x)
          i = i + 1

        keys2[j-4, i-1] = sum

        self.d2.append(sum)
        d1.append(self.d2)
        self.d2.clear()

      k = j
      i = 0

      for x in self.c:
        VAL = np.vstack((keys2[:,i],keys2[:,(len(self.c))]))
        R_xy = np.corrcoef(VAL)
        keys2[-1,i] = R_xy[1,0] #ЗАПОЛНЕНИЕ ВЕСОВ
          #необходимо удалять ключи, в которых отрицательная корреляция
        i = i + 1
      j = 0

  def matching_topic(self, message):
    #ФОРМИРОВАНИЕ СООБЩЕНИЙ, МАТРИЦЫ, И ФОРМУЛЫ

    # rb = xlrd.open_workbook('content/Messages.xls', formatting_info=True)
    # sheet = rb.sheet_by_index(0)

    self.mess3.clear()
    self.mess_one.clear()

    self.mess3.append(message.split())

    # k = 0
    # for rownum in range(sheet.nrows):
    #   row = sheet.row_values(rownum)
    #   for c_el in row:
    #     mess_one.extend(c_el.split())
    #   mess3.append(mess_one)
    #   k = k + 1
    #   mess_one = []

    for y in self.mess3:
      self.d3.append(Porter.stem_list(y))
    self.mess3.clear()
    self.mess3.extend(self.d3)

    wb = openpyxl.Workbook()
    wb.create_sheet(title = 'Test', index=0)
    sheet = wb['Test']

    keys3 = np.zeros([(len(self.mess3)+1),(len(self.c)+3)],float)

    i = 1
    for x in self.c:
      value = str(x)

      cell = sheet.cell(row=1, column=i) #ВЫВОД КЛЮЧЕЙ
      cell.value  = value

      i = i + 1

    self.d1.clear()
    self.d2.clear()
    j = 3
    for y in self.mess3:
      j = j + 1
      i = 1
      sum = 0
      for x in self.c:
        sum += y.count(x)
        value = y.count(x)
        cell = sheet.cell(row=j, column=i)
        cell.value = value
        keys3[j-4, i-1] = y.count(x)
        i = i + 1

      value = str(sum)
      cell = sheet.cell(row=j, column=i)
      cell.value = value

    # заполнение файла excel
    j = 3
    for y in self.mess3:
      j = j + 1
      i = 1
      sum = 0
      for x in self.c:
        sum += keys3[j-4, i-1]
        value = keys3[j-4, i-1]
        cell = sheet.cell(row=j, column=i)
        cell.value = value
        i = i + 1
      
      value = str(sum)
      cell = sheet.cell(row=j, column=i)
      cell.value = value

      self.d3.clear()

      return value #ЭТО ПРИЗНАК ТЕМЫ